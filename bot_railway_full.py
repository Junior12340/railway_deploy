import asyncio
import logging
import os
import re
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton,
    CallbackQuery, Message, FSInputFile
)
import aiosqlite
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

# ==================== SOZLAMALAR ====================
# Railway environment variables dan olinadi
BOT_TOKEN = os.getenv("BOT_TOKEN", "8168499228:AAHX-FgQc1Xnmgd7PuvPwauUQN4lQIPybPc")
GROUP_CHAT_ID = int(os.getenv("GROUP_CHAT_ID", "-1003773765959"))

# Kategoriyalar uchun guruh ID lari
CATEGORY_GROUPS = {
    "Yo'l va transport": int(os.getenv("GROUP_YOL_TRANSPORT", os.getenv("GROUP_CHAT_ID", "-1003773765959"))),
    "Kommunal xizmatlar": int(os.getenv("GROUP_KOMMUNAL", os.getenv("GROUP_CHAT_ID", "-1003773765959"))),
    "Ta'lim": int(os.getenv("GROUP_TALIM", os.getenv("GROUP_CHAT_ID", "-1003773765959"))),
    "Sog'liqni saqlash": int(os.getenv("GROUP_SOGLIQ", os.getenv("GROUP_CHAT_ID", "-1003773765959"))),
    "Ijtimoiy masalalar": int(os.getenv("GROUP_IJTIMOIY", os.getenv("GROUP_CHAT_ID", "-1003773765959"))),
    "Boshqa": int(os.getenv("GROUP_BOSHQA", os.getenv("GROUP_CHAT_ID", "-1003773765959")))
}

# Qo'shimcha sozlamalar
DAILY_LIMIT = int(os.getenv("DAILY_LIMIT", "1"))
REMINDER_DAYS = int(os.getenv("REMINDER_DAYS", "15"))
DB_PATH = os.getenv("DB_PATH", "murojaatlar.db")
MEDIA_PATH = os.getenv("MEDIA_PATH", "media_photos")
DEFAULT_IMAGE = "default_image.png"

# ==================== LOGGING ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== BOT VA DISPATCHER ====================
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# ==================== VALIDATSIYA FUNKSIYALARI ====================
def validate_passport(passport: str) -> bool:
    """Pasport tekshirish: AA1234567 (2 harf + 7 raqam)"""
    if not passport:
        return False
    pattern = r'^[A-Z]{2}\d{7}$'
    return bool(re.match(pattern, passport))

def validate_phone(phone: str) -> bool:
    """Telefon tekshirish: +998XXXXXXXXX yoki 998XXXXXXXXX"""
    if not phone:
        return False
    phone_clean = phone.replace('+', '').replace(' ', '').replace('-', '')
    pattern = r'^998\d{9}$'
    return bool(re.match(pattern, phone_clean))

def validate_full_name(full_name: str) -> bool:
    """F.I.Sh tekshirish: kamida 2 ta so'z, raqamsiz"""
    if not full_name:
        return False
    words = full_name.strip().split()
    if len(words) < 2:
        return False
    return all(not any(char.isdigit() for char in word) for word in words)

def get_target_group(category: str) -> int:
    """Kategoriyaga mos guruh ID ni qaytarish"""
    return CATEGORY_GROUPS.get(category, GROUP_CHAT_ID)

def get_all_group_ids() -> list:
    """Barcha guruh ID larini olish"""
    return list(set(CATEGORY_GROUPS.values()))

# ==================== MA'LUMOTLAR BAZASI ====================
class Database:
    """Database boshqaruvi"""
    
    def __init__(self):
        self.db_path = DB_PATH
        
    async def init_db(self):
        """Database yaratish va jadvallarni sozlash"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute("PRAGMA foreign_keys = OFF")
                
                # Foydalanuvchilar jadvali
                await db.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        user_id INTEGER PRIMARY KEY,
                        full_name TEXT,
                        phone TEXT,
                        welcome_shown INTEGER DEFAULT 0,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Murojaatlar jadvali
                await db.execute("""
                    CREATE TABLE IF NOT EXISTS murojaatlar (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        full_name TEXT,
                        passport TEXT,
                        phone TEXT,
                        address TEXT,
                        category TEXT,
                        text TEXT,
                        image_path TEXT,
                        status TEXT DEFAULT 'Yangi',
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        admin_checked_at DATETIME,
                        group_message_id INTEGER
                    )
                """)
                
                # Javoblar jadvali
                await db.execute("""
                    CREATE TABLE IF NOT EXISTS javoblar (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        murojaat_id INTEGER,
                        admin_id INTEGER,
                        admin_username TEXT,
                        javob_text TEXT,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Migratsiya: ustunlarni tekshirish va qo'shish
                try:
                    cursor = await db.execute("PRAGMA table_info(murojaatlar)")
                    columns = await cursor.fetchall()
                    column_names = [col[1] for col in columns]

                    required_columns = {
                        'address': 'TEXT',
                        'group_message_id': 'INTEGER',
                        'admin_checked_at': 'DATETIME'
                    }

                    for col_name, col_type in required_columns.items():
                        if col_name not in column_names:
                            logger.warning(f"⚠️ '{col_name}' ustuni topilmadi, qo'shilmoqda...")
                            await db.execute(f"ALTER TABLE murojaatlar ADD COLUMN {col_name} {col_type}")
                            await db.commit()
                            logger.info(f"✅ '{col_name}' ustuni muvaffaqiyatli qo'shildi")

                    cursor = await db.execute("PRAGMA table_info(users)")
                    columns = await cursor.fetchall()
                    column_names = [col[1] for col in columns]

                    if 'welcome_shown' not in column_names:
                        logger.warning(f"⚠️ 'welcome_shown' ustuni topilmadi, qo'shilmoqda...")
                        await db.execute("ALTER TABLE users ADD COLUMN welcome_shown INTEGER DEFAULT 0")
                        await db.commit()
                        logger.info(f"✅ 'welcome_shown' ustuni muvaffaqiyatli qo'shildi")

                    logger.info("✅ Barcha ustunlar tekshirildi")

                except Exception as migration_error:
                    logger.error(f"❌ Migratsiya xatolik: {migration_error}")
                
                await db.commit()
                logger.info("✅ Database tayyor")
        except Exception as e:
            logger.error(f"❌ Database xatolik: {e}")
            raise
    
    async def add_user(self, user_id: int, full_name: str, phone: str):
        """Foydalanuvchi qo'shish"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute("PRAGMA foreign_keys = OFF")
                await db.execute(
                    "INSERT OR REPLACE INTO users (user_id, full_name, phone) VALUES (?, ?, ?)",
                    (user_id, full_name, phone)
                )
                await db.commit()
                logger.info(f"✅ User qo'shildi: {user_id}")
        except Exception as e:
            logger.error(f"❌ User qo'shish xatolik: {e}")

    async def user_exists(self, user_id: int) -> bool:
        """Foydalanuvchi mavjud ekanligini tekshirish"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                async with db.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,)) as cursor:
                    return await cursor.fetchone() is not None
        except Exception as e:
            logger.error(f"❌ User check xatolik: {e}")
            return False

    async def is_welcome_shown(self, user_id: int) -> bool:
        """Eslatma ko'rsatilganligini tekshirish"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                async with db.execute("SELECT welcome_shown FROM users WHERE user_id = ?", (user_id,)) as cursor:
                    row = await cursor.fetchone()
                    return row and row[0] == 1
        except Exception as e:
            logger.error(f"❌ Welcome check xatolik: {e}")
            return False

    async def mark_welcome_shown(self, user_id: int):
        """Eslatmani ko'rsatilgan deb belgilash"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute("UPDATE users SET welcome_shown = 1 WHERE user_id = ?", (user_id,))
                await db.commit()
                logger.info(f"✅ Welcome marked: {user_id}")
        except Exception as e:
            logger.error(f"❌ Mark welcome xatolik: {e}")
    
    async def add_murojaat(self, user_id: int, full_name: str, passport: str, 
                          phone: str, address: str, category: str, text: str, 
                          image_path: str = None, group_message_id: int = None):
        """Murojaat qo'shish"""
        try:
            logger.info(f"💾 Murojaat saqlanmoqda: user={user_id}, group_msg={group_message_id}")
            
            await self.add_user(user_id, full_name, phone)
            
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute("PRAGMA foreign_keys = OFF")
                cursor = await db.execute("""
                    INSERT INTO murojaatlar 
                    (user_id, full_name, passport, phone, address, category, text, image_path, group_message_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (user_id, full_name, passport, phone, address, category, text, image_path, group_message_id))
                await db.commit()
                murojaat_id = cursor.lastrowid
                
                logger.info(f"✅ Murojaat saqlandi: ID={murojaat_id}")
                return murojaat_id
                
        except Exception as e:
            logger.error(f"❌ Murojaat qo'shish xatolik: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    async def get_user_murojaatlar(self, user_id: int):
        """Foydalanuvchi murojaatlari"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM murojaatlar WHERE user_id = ? ORDER BY created_at DESC",
                    (user_id,)
                ) as cursor:
                    rows = await cursor.fetchall()
                    return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"❌ Get user murojaatlar xatolik: {e}")
            return []
    
    async def get_murojaat_javoblar(self, murojaat_id: int):
        """Murojaat javoblari"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM javoblar WHERE murojaat_id = ? ORDER BY created_at DESC",
                    (murojaat_id,)
                ) as cursor:
                    rows = await cursor.fetchall()
                    return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"❌ Get javoblar xatolik: {e}")
            return []
    
    async def get_murojaat_by_group_msg(self, group_message_id: int):
        """Guruh xabari bo'yicha murojaatni topish"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM murojaatlar WHERE group_message_id = ?",
                    (group_message_id,)
                ) as cursor:
                    row = await cursor.fetchone()
                    return dict(row) if row else None
        except Exception as e:
            logger.error(f"❌ Get murojaat by group msg xatolik: {e}")
            return None
    
    async def add_javob(self, murojaat_id: int, admin_id: int, admin_username: str, javob_text: str):
        """Javob qo'shish"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute("""
                    INSERT INTO javoblar (murojaat_id, admin_id, admin_username, javob_text)
                    VALUES (?, ?, ?, ?)
                """, (murojaat_id, admin_id, admin_username, javob_text))
                await db.commit()
                logger.info(f"✅ Javob saqlandi: murojaat_id={murojaat_id}")
        except Exception as e:
            logger.error(f"❌ Javob qo'shish xatolik: {e}")
    
    async def update_status(self, murojaat_id: int, status: str):
        """Status yangilash"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                await db.execute(
                    "UPDATE murojaatlar SET status = ?, admin_checked_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (status, murojaat_id)
                )
                await db.commit()
                logger.info(f"✅ Status yangilandi: #{murojaat_id} -> {status}")
        except Exception as e:
            logger.error(f"❌ Status yangilash xatolik: {e}")
    
    async def get_daily_count(self, user_id: int):
        """Bugungi murojaatlar soni"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                today = datetime.now().strftime('%Y-%m-%d')
                async with db.execute(
                    "SELECT COUNT(*) FROM murojaatlar WHERE user_id = ? AND DATE(created_at) = ?",
                    (user_id, today)
                ) as cursor:
                    row = await cursor.fetchone()
                    return row[0] if row else 0
        except Exception as e:
            logger.error(f"❌ Daily count xatolik: {e}")
            return 0
    
    async def get_pending_murojaatlar(self):
        """Javob kutayotgan murojaatlar"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM murojaatlar WHERE status = 'Yangi' ORDER BY created_at ASC"
                ) as cursor:
                    rows = await cursor.fetchall()
                    return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"❌ Get pending xatolik: {e}")
            return []
    
    async def get_all_statistics(self):
        """To'liq statistika"""
        try:
            async with aiosqlite.connect(self.db_path) as db:
                db.row_factory = aiosqlite.Row
                
                # Umumiy soni
                async with db.execute("SELECT COUNT(*) as total FROM murojaatlar") as cursor:
                    total = (await cursor.fetchone())['total']
                
                # Javob berilgan
                async with db.execute("SELECT COUNT(*) FROM murojaatlar WHERE status = 'Javob berildi'") as cursor:
                    answered = (await cursor.fetchone())[0]
                
                # Javob kutayotgan
                async with db.execute("SELECT COUNT(*) FROM murojaatlar WHERE status = 'Yangi'") as cursor:
                    pending = (await cursor.fetchone())[0]
                
                # Bugungi
                today = datetime.now().strftime('%Y-%m-%d')
                async with db.execute(
                    "SELECT COUNT(*) FROM murojaatlar WHERE DATE(created_at) = ?",
                    (today,)
                ) as cursor:
                    today_count = (await cursor.fetchone())[0]
                
                # Haftalik
                week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
                async with db.execute(
                    "SELECT COUNT(*) FROM murojaatlar WHERE DATE(created_at) >= ?",
                    (week_ago,)
                ) as cursor:
                    weekly = (await cursor.fetchone())[0]
                
                # Kategoriyalar bo'yicha
                async with db.execute("""
                    SELECT category, COUNT(*) as count 
                    FROM murojaatlar 
                    GROUP BY category 
                    ORDER BY count DESC
                """) as cursor:
                    categories = [dict(row) for row in await cursor.fetchall()]
                
                return {
                    'total': total,
                    'answered': answered,
                    'pending': pending,
                    'today': today_count,
                    'weekly': weekly,
                    'categories': categories
                }
        except Exception as e:
            logger.error(f"❌ Statistika xatolik: {e}")
            return None

# Database instance
db = Database()

# ==================== FSM STATES ====================
class MurojaatStates(StatesGroup):
    """Murojaat yuborish holatlari"""
    full_name = State()
    passport = State()
    phone = State()
    address = State()
    category = State()
    text = State()
    photo = State()

# ==================== KEYBOARD ====================
def get_main_menu():
    """Asosiy menyu klaviaturasi"""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📝 Murojaat yuborish")],
            [KeyboardButton(text="📋 Mening murojaatlarim")],
            [KeyboardButton(text="ℹ️ Ma'lumot"), KeyboardButton(text="📞 Aloqa")]
        ],
        resize_keyboard=True
    )
    return keyboard

def get_categories_keyboard():
    """Kategoriyalar klaviaturasi"""
    categories = [
        "Yo'l va transport",
        "Kommunal xizmatlar",
        "Ta'lim",
        "Sog'liqni saqlash",
        "Ijtimoiy masalalar",
        "Boshqa"
    ]
    
    keyboard = [[KeyboardButton(text=cat)] for cat in categories]
    keyboard.append([KeyboardButton(text="🔙 Bekor qilish")])
    
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)

def get_photo_keyboard():
    """Rasm yuklash klaviaturasi"""
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📸 Rasm yuklash", callback_data="add_photo")],
            [InlineKeyboardButton(text="⏭ Rasmsiz davom etish", callback_data="skip_photo")]
        ]
    )
    return keyboard

# ==================== REMINDER SCHEDULER ====================
class ReminderScheduler:
    """Eslatmalar rejasi"""
    
    def __init__(self, bot: Bot):
        self.bot = bot
        self.scheduler = AsyncIOScheduler()
    
    def start(self):
        """Schedulerni ishga tushirish"""
        self.scheduler.add_job(
            self.send_reminders,
            'cron',
            hour=10,
            minute=0,
            id='reminder_job'
        )
        self.scheduler.start()
        logger.info("✅ Reminder scheduler ishga tushdi")
    
    async def send_reminders(self):
        """Eslatmalarni yuborish"""
        try:
            cutoff_date = (datetime.now() - timedelta(days=REMINDER_DAYS)).strftime('%Y-%m-%d')
            
            async with aiosqlite.connect(DB_PATH) as db_conn:
                db_conn.row_factory = aiosqlite.Row
                async with db_conn.execute(
                    """SELECT * FROM murojaatlar 
                       WHERE status = 'Yangi' 
                       AND DATE(created_at) <= ?""",
                    (cutoff_date,)
                ) as cursor:
                    old_requests = await cursor.fetchall()
            
            if old_requests:
                message = (
                    f"⚠️ <b>ESLATMA!</b>\n\n"
                    f"Javob kutayotgan eski murojaatlar: <b>{len(old_requests)} ta</b>\n"
                    f"({REMINDER_DAYS} kundan oshgan)\n\n"
                    f"Iltimos, ko'rib chiqing!"
                )

                for group_id in get_all_group_ids():
                    try:
                        await self.bot.send_message(group_id, message, parse_mode="HTML")
                    except Exception as group_error:
                        logger.error(f"❌ Guruhga eslatma yuborish xatolik {group_id}: {group_error}")

                logger.info(f"📨 Eslatma yuborildi: {len(old_requests)} ta eski murojaat")
        
        except Exception as e:
            logger.error(f"❌ Reminder xatolik: {e}")

# ==================== BOT HANDLERS ====================
@dp.message(CommandStart())
async def cmd_start(message: Message):
    """Start komandasi"""
    user_id = message.from_user.id
    user_exists = await db.user_exists(user_id)
    welcome_already_shown = await db.is_welcome_shown(user_id)

    if not user_exists or not welcome_already_shown:
        disclaimer_text = (
            "⚠️ <b>DIQQAT!</b>\n\n"
            "📌 Bu telegram bot <b>Qashqadaryo viloyati Dehqonobod tumani Hokimyati</b>ga tegishli.\n\n"
            "ℹ️ Bot orqali siz:\n"
            "• Turli masalalar bo'yicha murojaat yuborishingiz mumkin\n"
            "• Murojaatlaringiz holatini kuzatishingiz mumkin\n"
            "• Javoblar olishingiz mumkin\n\n"
            "✅ Botdan foydalanishni davom etish uchun pastdagi tugmani bosing 👇"
        )
        await message.answer(disclaimer_text, parse_mode="HTML")
        await asyncio.sleep(1)
        await db.mark_welcome_shown(user_id)

    welcome_text = (
        f"👋 <b>Assalomu alaykum, {message.from_user.first_name}!</b>\n\n"
        "📝 Men Qashqadaryo viloyati Dehqonobod tumani murojaatlar botiman. Men orqali siz:\n"
        "• Turli masalalar bo'yicha murojaat yuborishingiz\n"
        "• Murojaatlaringiz holatini kuzatishingiz\n"
        "• Javoblarni olishingiz mumkin\n\n"
        "Boshlash uchun quyidagi tugmalardan birini tanlang 👇"
    )

    await message.answer(welcome_text, reply_markup=get_main_menu(), parse_mode="HTML")

@dp.message(F.text == "📝 Murojaat yuborish")
async def start_murojaat(message: Message, state: FSMContext):
    """Murojaat yuborishni boshlash"""
    # Kunlik limitni tekshirish
    daily_count = await db.get_daily_count(message.from_user.id)
    
    if daily_count >= DAILY_LIMIT:
        await message.answer(
            f"⚠️ <b>Kunlik limit tugadi!</b>\n\n"
            f"Siz bugun allaqachon {daily_count} ta murojaat yuborgansiz.\n"
            f"Maksimal ruxsat: {DAILY_LIMIT} ta murojaat 1 kunda\n\n"
            f"Iltimos, ertaga qayta urinib ko'ring.",
            parse_mode="HTML"
        )
        return
    
    await message.answer(
        "👤 <b>To'liq ismingizni kiriting</b>\n\n"
        "Masalan: Aliyev Vali Valiyevich",
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.full_name)

@dp.message(MurojaatStates.full_name)
async def process_full_name(message: Message, state: FSMContext):
    """F.I.Sh qabul qilish"""
    full_name = message.text.strip()
    
    if not validate_full_name(full_name):
        await message.answer(
            "❌ Noto'g'ri format!\n\n"
            "To'liq ismingizni kiriting (kamida 2 ta so'z, raqamsiz)\n"
            "Masalan: Aliyev Vali Valiyevich"
        )
        return
    
    await state.update_data(full_name=full_name)
    await message.answer(
        "🛂 <b>Pasport seriya va raqamini kiriting</b>\n\n"
        "Format: AA1234567 (2 ta harf + 7 ta raqam)",
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.passport)

@dp.message(MurojaatStates.passport)
async def process_passport(message: Message, state: FSMContext):
    """Pasport qabul qilish"""
    passport = message.text.strip().upper()
    
    if not validate_passport(passport):
        await message.answer(
            "❌ Noto'g'ri format!\n\n"
            "To'g'ri format: AA1234567\n"
            "(2 ta harf + 7 ta raqam)"
        )
        return
    
    await state.update_data(passport=passport)
    await message.answer(
        "📱 <b>Telefon raqamingizni kiriting</b>\n\n"
        "Format: +998XXXXXXXXX",
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.phone)

@dp.message(MurojaatStates.phone)
async def process_phone(message: Message, state: FSMContext):
    """Telefon qabul qilish"""
    phone = message.text.strip()
    
    if not validate_phone(phone):
        await message.answer(
            "❌ Noto'g'ri format!\n\n"
            "To'g'ri format: +998XXXXXXXXX"
        )
        return
    
    await state.update_data(phone=phone)
    await message.answer(
        "🏠 <b>Manzilingizni kiriting</b>\n\n"
        "Masalan: Qarashina shahri, Istiqlol M.F.Y",
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.address)

@dp.message(MurojaatStates.address)
async def process_address(message: Message, state: FSMContext):
    """Manzil qabul qilish"""
    address = message.text.strip()
    
    if len(address) < 5:
        await message.answer("❌ Manzil juda qisqa! Kamida 5 ta belgi kiriting.")
        return
    
    await state.update_data(address=address)
    await message.answer(
        "📂 <b>Murojaat turini tanlang:</b>",
        reply_markup=get_categories_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.category)

@dp.message(MurojaatStates.category)
async def process_category(message: Message, state: FSMContext):
    """Kategoriya qabul qilish"""
    if message.text == "🔙 Bekor qilish":
        await message.answer(
            "❌ Murojaat bekor qilindi.",
            reply_markup=get_main_menu()
        )
        await state.clear()
        return
    
    category = message.text.strip()
    await state.update_data(category=category)
    
    await message.answer(
        "📝 <b>Murojaatingiz matnini kiriting</b>\n\n"
        "Batafsil yozing (kamida 10 ta belgi):",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Bekor qilish")]],
            resize_keyboard=True
        ),
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.text)

@dp.message(MurojaatStates.text)
async def process_text(message: Message, state: FSMContext):
    """Matn qabul qilish"""
    if message.text == "🔙 Bekor qilish":
        await message.answer(
            "❌ Murojaat bekor qilindi.",
            reply_markup=get_main_menu()
        )
        await state.clear()
        return
    
    text = message.text.strip()
    
    if len(text) < 10:
        await message.answer("❌ Matn juda qisqa! Kamida 10 ta belgi kiriting.")
        return
    
    await state.update_data(text=text)
    await message.answer(
        "📸 <b>Rasm yuklashni xohlaysizmi?</b>\n\n"
        "Rasm murojaat bilan birga yuboriladi.",
        reply_markup=get_photo_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(MurojaatStates.photo)

@dp.callback_query(F.data == "add_photo")
async def add_photo_callback(callback: CallbackQuery):
    """Rasm yuklash tugmasi"""
    await callback.message.edit_text(
        "📸 <b>Rasmni yuboring</b>\n\n"
        "Faqat bitta rasm yuboring.",
        parse_mode="HTML"
    )
    await callback.answer()

@dp.callback_query(F.data == "skip_photo")
async def skip_photo_callback(callback: CallbackQuery, state: FSMContext):
    """Rasmsiz davom etish"""
    await callback.answer()
    await finish_murojaat(
        callback.message, 
        state, 
        photo_path=None,
        user_id=callback.from_user.id
    )

@dp.message(MurojaatStates.photo, F.photo)
async def process_photo(message: Message, state: FSMContext):
    """Rasmni qabul qilish"""
    photo = message.photo[-1]
    
    file = await bot.get_file(photo.file_id)
    file_extension = file.file_path.split('.')[-1]
    filename = f"{message.from_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{file_extension}"
    photo_path = os.path.join(MEDIA_PATH, filename)
    
    await bot.download_file(file.file_path, photo_path)
    logger.info(f"✅ Rasm saqlandi: {photo_path}")
    
    await finish_murojaat(message, state, photo_path=photo_path)

async def finish_murojaat(message: Message, state: FSMContext, photo_path: str = None, user_id: int = None):
    """Murojaatni yakunlash"""
    data = await state.get_data()
    
    required_fields = ['full_name', 'passport', 'phone', 'address', 'category', 'text']
    if not all(field in data for field in required_fields):
        await message.answer(
            "❌ Xatolik: Ma'lumotlar to'liq emas.",
            reply_markup=get_main_menu()
        )
        await state.clear()
        return
    
    confirm_text = (
        "✅ <b>MUROJAAT TAYYORLANDI</b>\n\n"
        f"👤 <b>F.I.Sh:</b> {data['full_name']}\n"
        f"🛂 <b>Pasport:</b> {data['passport']}\n"
        f"📱 <b>Telefon:</b> {data['phone']}\n"
        f"🏠 <b>Manzil:</b> {data['address']}\n"
        f"📂 <b>Tur:</b> {data['category']}\n"
        f"📝 <b>Matn:</b> {data['text']}\n"
        f"📸 <b>Rasm:</b> {'✅ Bor' if photo_path else '❌ Yoq'}\n\n"
        "<i>Murojaat guruhga yuborilmoqda...</i>"
    )
    
    await message.answer(confirm_text, parse_mode="HTML")

    try:
        target_group_id = get_target_group(data['category'])

        group_text = (
            "🆕 <b>YANGI MUROJAAT</b>\n\n"
            f"👤 <b>F.I.Sh:</b> {data['full_name']}\n"
            f"🛂 <b>Pasport:</b> {data['passport']}\n"
            f"📱 <b>Telefon:</b> {data['phone']}\n"
            f"🏠 <b>Manzil:</b> {data['address']}\n"
            f"📂 <b>Tur:</b> {data['category']}\n"
            f"📝 <b>Matn:</b>\n{data['text']}\n\n"
            f"📅 <b>Sana:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
            f"🔢 <b>User ID:</b> <code>{message.from_user.id}</code>\n\n"
            "<i>💬 Javob berish uchun bu xabarga reply qiling!</i>"
        )

        if photo_path and os.path.exists(photo_path):
            photo_file = FSInputFile(photo_path)
            sent_message = await bot.send_photo(
                target_group_id,
                photo=photo_file,
                caption=group_text,
                parse_mode="HTML"
            )
            final_image_path = photo_path
        else:
            if os.path.exists(DEFAULT_IMAGE):
                default_photo = FSInputFile(DEFAULT_IMAGE)
                sent_message = await bot.send_photo(
                    target_group_id,
                    photo=default_photo,
                    caption=group_text,
                    parse_mode="HTML"
                )
                final_image_path = DEFAULT_IMAGE
            else:
                sent_message = await bot.send_message(
                    target_group_id,
                    group_text,
                    parse_mode="HTML"
                )
                final_image_path = None
        
        group_message_id = sent_message.message_id
        logger.info(f"✅ Guruhga yuborildi: message_id={group_message_id}")
        
        actual_user_id = user_id if user_id else message.from_user.id
        
        murojaat_id = await db.add_murojaat(
            user_id=actual_user_id,
            full_name=data['full_name'],
            passport=data['passport'],
            phone=data['phone'],
            address=data['address'],
            category=data['category'],
            text=data['text'],
            image_path=final_image_path,
            group_message_id=group_message_id
        )
        
        if murojaat_id:
            success_text = (
                "✅ <b>MUROJAAT YUBORILDI!</b>\n\n"
                f"📋 Murojaat raqami: <b>#{murojaat_id}</b>\n\n"
                "📬 Murojaatingiz ko'rib chiqilmoqda.\n"
                "Javob kelganda xabar beramiz.\n\n"
                "📊 Holatni \"📋 Mening murojaatlarim\" orqali kuzating."
            )
            await message.answer(success_text, reply_markup=get_main_menu(), parse_mode="HTML")
        else:
            await message.answer(
                "⚠️ Murojaat yuborildi, lekin saqlashda xatolik.",
                reply_markup=get_main_menu()
            )
        
    except Exception as e:
        logger.error(f"❌ Guruhga yuborish xatolik: {e}")
        import traceback
        traceback.print_exc()
        await message.answer(
            f"❌ <b>Xatolik!</b>\n\n"
            f"Murojaat yuborib bo'lmadi.\n\n"
            f"<code>{str(e)}</code>",
            reply_markup=get_main_menu(),
            parse_mode="HTML"
        )
    
    await state.clear()

# ==================== GURUHDA JAVOB BERISH ====================
@dp.message(F.reply_to_message)
async def group_reply_handler(message: Message):
    """Guruhda javob berish"""
    try:
        if not message.reply_to_message:
            return

        if message.chat.id not in get_all_group_ids():
            return
        
        reply_to_message_id = message.reply_to_message.message_id
        logger.info(f"🔍 Guruhda javob: reply_to={reply_to_message_id}")
        
        murojaat = await db.get_murojaat_by_group_msg(reply_to_message_id)
        
        if not murojaat:
            logger.warning(f"⚠️ Murojaat topilmadi: {reply_to_message_id}")
            await message.reply(
                "❌ <b>Murojaat topilmadi!</b>\n\n"
                f"Reply ID: <code>{reply_to_message_id}</code>\n\n"
                "Murojaat xabariga to'g'ridan-to'g'ri reply qiling.",
                parse_mode="HTML"
            )
            return
        
        murojaat_id = murojaat['id']
        user_id = murojaat['user_id']
        javob_text = message.text
        
        logger.info(f"✅ Murojaat topildi: #{murojaat_id}, user={user_id}")
        
        if not javob_text or len(javob_text.strip()) < 3:
            await message.reply("❌ Javob juda qisqa! Kamida 3 ta belgi.")
            return
        
        admin_id = message.from_user.id
        admin_username = message.from_user.username or message.from_user.first_name or f"Admin{admin_id}"
        
        await db.add_javob(murojaat_id, admin_id, admin_username, javob_text)
        await db.update_status(murojaat_id, "Javob berildi")
        
        await message.reply(
            f"✅ <b>JAVOB YUBORILDI!</b>\n\n"
            f"📋 Murojaat: #{murojaat_id}\n"
            f"👤 Admin: @{admin_username}\n"
            f"💬 Javob: {javob_text[:100]}{'...' if len(javob_text) > 100 else ''}\n\n"
            f"<i>✓ Javob yuborildi</i>",
            parse_mode="HTML"
        )
        
        try:
            bot_info = await bot.get_me()
            
            if user_id == bot_info.id:
                logger.error(f"❌ User ID bot IDsi!")
                await message.reply(
                    f"❌ <b>DATABASE XATOLIGI!</b>\n\n"
                    f"User ID bot IDsi: <code>{user_id}</code>\n"
                    f"Yangi murojaat yuborib ko'ring.",
                    parse_mode="HTML"
                )
                return
            
            await bot.send_message(
                user_id,
                f"📬 <b>#{murojaat_id} raqamli murojaatingizga javob!</b>\n\n"
                f"💬 <b>Javob:</b>\n{javob_text}\n\n"
                f"Rahmat! 🙏",
                parse_mode="HTML"
            )
            logger.info(f"✅ Javob yuborildi: user={user_id}")
        except Exception as e:
            error_message = str(e)
            logger.error(f"❌ Foydalanuvchiga yuborish xatolik: {error_message}")
            
            if "bots can't send messages to bots" in error_message:
                error_text = "Bot tomonidan yuborilgan."
            elif "bot was blocked by the user" in error_message:
                error_text = "Foydalanuvchi botni bloklagan."
            elif "user is deactivated" in error_message:
                error_text = "Akkaunt o'chirilgan."
            elif "chat not found" in error_message:
                error_text = "Foydalanuvchi /start qilmagan."
            else:
                error_text = f"Xatolik: {error_message}"
            
            await message.reply(
                f"⚠️ <b>Javob saqlandi, lekin yuborilmadi</b>\n\n"
                f"📋 #{murojaat_id}\n"
                f"👤 User: <code>{user_id}</code>\n"
                f"❌ {error_text}",
                parse_mode="HTML"
            )
    
    except Exception as e:
        logger.error(f"❌ Reply handler xatolik: {e}")
        import traceback
        traceback.print_exc()

# ==================== MUROJAATLARIM ====================
@dp.message(F.text == "📋 Mening murojaatlarim")
async def my_murojaatlar(message: Message):
    """Foydalanuvchi murojaatlari"""
    murojaatlar = await db.get_user_murojaatlar(message.from_user.id)
    
    if not murojaatlar:
        await message.answer(
            "📭 <b>Sizda hali murojaatlar yo'q</b>\n\n"
            "Murojaat yuborish uchun \"📝 Murojaat yuborish\" tugmasini bosing.",
            parse_mode="HTML"
        )
        return
    
    response = f"📋 <b>MUROJAATLARIM</b>\n\n"
    response += f"Jami: {len(murojaatlar)} ta\n\n"
    
    for m in murojaatlar[:10]:
        javoblar = await db.get_murojaat_javoblar(m['id'])
        status_emoji = "✅" if m['status'] == "Javob berildi" else "⏳"
        
        response += (
            f"{status_emoji} <b>#{m['id']}</b> - {m['category']}\n"
            f"📅 {m['created_at'][:16]}\n"
            f"📊 Status: {m['status']}\n"
        )
        
        if javoblar:
            response += f"💬 Javoblar: {len(javoblar)} ta\n"
            for j in javoblar[:1]:
                response += f"   └ {j['javob_text'][:50]}...\n"
        
        response += "\n"
    
    if len(murojaatlar) > 10:
        response += f"<i>... va yana {len(murojaatlar) - 10} ta</i>\n"
    
    await message.answer(response, parse_mode="HTML")

# ==================== MA'LUMOT VA ALOQA ====================
@dp.message(F.text == "ℹ️ Ma'lumot")
async def info_handler(message: Message):
    """Ma'lumot"""
    info_text = (
        "ℹ️ <b>BOT HAQIDA</b>\n\n"
        "Bu bot orqali siz turli masalalar bo'yicha murojaat yuborishingiz mumkin.\n\n"
        f"📊 <b>Limitlar:</b>\n"
        f"• Kunlik: {DAILY_LIMIT} ta murojaat\n"
        f"• Eslatma: {REMINDER_DAYS} kun\n\n"
        "📂 <b>Kategoriyalar:</b>\n"
        "• Yo'l va transport\n"
        "• Kommunal xizmatlar\n"
        "• Ta'lim\n"
        "• Sog'liqni saqlash\n"
        "• Ijtimoiy masalalar\n"
        "• Boshqa\n\n"
        "✅ Barcha murojaatlar ko'rib chiqiladi!"
    )
    await message.answer(info_text, parse_mode="HTML")

@dp.message(F.text == "📞 Aloqa")
async def contact_handler(message: Message):
    """Aloqa"""
    contact_text = (
        "📞 <b>ALOQA</b>\n\n"
        "Savollar bo'lsa:\n"
        "📧 Email: support@example.com\n"
        "📱 Telefon: +998 XX XXX XX XX\n\n"
        "Yoki admin bilan bog'laning."
    )
    await message.answer(contact_text, parse_mode="HTML")

# ==================== EXCEL EXPORT ====================
async def create_excel_report():
    """Excel hisobot yaratish"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistika"
        
        # Statistika
        stats = await db.get_all_statistics()
        
        ws['A1'] = "MUROJAATLAR STATISTIKASI"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Ko'rsatkich"
        ws['B3'] = "Qiymat"
        
        ws['A4'] = "Jami murojaatlar"
        ws['B4'] = stats['total']
        
        ws['A5'] = "Javob berilgan"
        ws['B5'] = stats['answered']
        
        ws['A6'] = "Javob kutayotgan"
        ws['B6'] = stats['pending']
        
        ws['A7'] = "Bugun"
        ws['B7'] = stats['today']
        
        ws['A8'] = "Oxirgi 7 kun"
        ws['B8'] = stats['weekly']
        
        # Kategoriyalar
        ws['A10'] = "KATEGORIYALAR"
        ws['A10'].font = Font(bold=True, size=12)
        
        row = 11
        for cat in stats['categories']:
            ws[f'A{row}'] = cat['category']
            ws[f'B{row}'] = cat['count']
            row += 1
        
        # Murojaatlar
        ws2 = wb.create_sheet("Murojaatlar")
        headers = ['ID', 'Sana', 'F.I.Sh', 'Telefon', 'Kategoriya', 'Status', 'Javoblar']
        ws2.append(headers)
        
        async with aiosqlite.connect(DB_PATH) as db_conn:
            db_conn.row_factory = aiosqlite.Row
            async with db_conn.execute(
                "SELECT * FROM murojaatlar ORDER BY created_at DESC"
            ) as cursor:
                murojaatlar = await cursor.fetchall()
        
        for m in murojaatlar:
            javoblar = await db.get_murojaat_javoblar(m['id'])
            javob_count = len(javoblar)
            
            ws2.append([
                m['id'],
                m['created_at'][:16],
                m['full_name'],
                m['phone'],
                m['category'],
                m['status'],
                javob_count
            ])
        
        # Styling
        for ws in [wb['Statistika'], wb['Murojaatlar']]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
        
        filename = f"murojaatlar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(MEDIA_PATH, filename)
        
        wb.save(filepath)
        logger.info(f"✅ Excel yaratildi: {filepath}")
        
        return filepath
    
    except Exception as e:
        logger.error(f"❌ Excel yaratish xatolik: {e}")
        import traceback
        traceback.print_exc()
        return None

# ==================== GURUH KOMANDALAR ====================
@dp.message(Command("stats"))
async def cmd_stats(message: Message):
    """Statistika - faqat guruhda"""
    if message.chat.id not in get_all_group_ids():
        if message.chat.type == "private":
            await message.answer("❌ Bu komanda faqat guruhda ishlaydi!")
        return
    
    try:
        stats = await db.get_all_statistics()
        
        if not stats:
            await message.answer("❌ Statistika topilmadi.")
            return
        
        categories_text = ""
        for cat in stats['categories']:
            categories_text += f"   • {cat['category']}: {cat['count']} ta\n"
        
        answered_percent = (stats['answered'] / stats['total'] * 100) if stats['total'] > 0 else 0
        pending_percent = (stats['pending'] / stats['total'] * 100) if stats['total'] > 0 else 0
        
        response = (
            "📊 <b>MUROJAATLAR STATISTIKASI</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            f"📈 <b>UMUMIY:</b>\n"
            f"   • Jami: <b>{stats['total']} ta</b>\n"
            f"   • Javob berilgan: <b>{stats['answered']} ta</b> ({answered_percent:.1f}%)\n"
            f"   • Javob kutayotgan: <b>{stats['pending']} ta</b> ({pending_percent:.1f}%)\n\n"
            
            f"📅 <b>DAVR:</b>\n"
            f"   • Bugun: <b>{stats['today']} ta</b>\n"
            f"   • 7 kun: <b>{stats['weekly']} ta</b>\n\n"
            
            f"📂 <b>KATEGORIYALAR:</b>\n"
            f"{categories_text}\n"
            
            "━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        
        await message.answer(response, parse_mode="HTML")
        
    except Exception as e:
        logger.error(f"❌ Statistika xatolik: {e}")
        await message.answer(f"❌ Xatolik: {e}")

@dp.message(Command("export"))
async def cmd_export(message: Message):
    """Excel export - faqat guruhda"""
    if message.chat.id not in get_all_group_ids():
        if message.chat.type == "private":
            await message.answer("❌ Bu komanda faqat guruhda ishlaydi!")
        return
    
    try:
        wait_msg = await message.answer("📊 Excel yaratilmoqda...")
        
        excel_path = await create_excel_report()
        
        if not excel_path or not os.path.exists(excel_path):
            await wait_msg.edit_text("❌ Excel yaratish xatolik.")
            return
        
        excel_file = FSInputFile(excel_path)
        await message.answer_document(
            document=excel_file,
            caption=(
                "📊 <b>MUROJAATLAR HISOBOTI</b>\n\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                f"📁 {os.path.basename(excel_path)}\n\n"
                "📈 Faylda:\n"
                "   • To'liq statistika\n"
                "   • Barcha murojaatlar\n"
                "   • Kategoriyalar\n\n"
                "<i>Excel da ko'ring!</i>"
            ),
            parse_mode="HTML"
        )
        
        await wait_msg.delete()
        logger.info(f"✅ Excel yuborildi: {excel_path}")
        
        try:
            os.remove(excel_path)
        except:
            pass
        
    except Exception as e:
        logger.error(f"❌ Export xatolik: {e}")
        await message.answer(f"❌ Xatolik: {e}")

@dp.message(Command("debug"))
async def cmd_debug(message: Message):
    """Debug"""
    try:
        async with aiosqlite.connect(DB_PATH) as db_conn:
            db_conn.row_factory = aiosqlite.Row
            async with db_conn.execute(
                "SELECT id, user_id, full_name, group_message_id, status, created_at FROM murojaatlar ORDER BY id DESC LIMIT 5"
            ) as cursor:
                rows = await cursor.fetchall()
        
        if not rows:
            await message.answer("📭 Database bo'sh!")
            return
        
        response = "🔍 <b>OXIRGI 5 TA:</b>\n\n"
        for r in rows:
            response += (
                f"📋 ID: {r['id']}\n"
                f"👤 User: {r['user_id']}\n"
                f"📛 {r['full_name']}\n"
                f"💬 Group: {r['group_message_id']}\n"
                f"📊 {r['status']}\n"
                f"📅 {r['created_at'][:16]}\n"
                f"━━━━━━━━━━\n"
            )
        
        await message.answer(response, parse_mode="HTML")
    
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")

# ==================== MAIN ====================
async def main():
    """Asosiy funksiya"""
    try:
        os.makedirs(MEDIA_PATH, exist_ok=True)
        logger.info(f"✅ Media papka: {MEDIA_PATH}")
        
        await db.init_db()
        logger.info("✅ Database tayyor")
        
        scheduler = ReminderScheduler(bot)
        scheduler.start()
        logger.info("✅ Scheduler tayyor")

        logger.info("🤖 Bot ishga tushmoqda...")
        logger.info(f"📊 Limit: {DAILY_LIMIT}/kun")
        logger.info(f"⏰ Eslatma: {REMINDER_DAYS} kun")
        logger.info(f"👥 Default Guruh: {GROUP_CHAT_ID}")
        logger.info("📂 Kategoriya guruhlari:")
        for category, group_id in CATEGORY_GROUPS.items():
            logger.info(f"   - {category}: {group_id}")
        logger.info("✅ Bot ishga tushdi!")
        
        await dp.start_polling(bot)
        
    except Exception as e:
        logger.error(f"❌ Bot xatolik: {e}")
        import traceback
        traceback.print_exc()
    finally:
        await bot.session.close()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("⏹ Bot to'xtatildi")
    except Exception as e:
        logger.error(f"❌ Xatolik: {e}")